#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Need to Install on the Anaconda Prompt:
    $ pip install pyexcel
"""


# Basic Modules
import os
import sys
# Read/Write to Excel
import csv
import pyexcel
import openpyxl as xl
# Openpyxl Styles
from openpyxl.styles import Alignment
from openpyxl.styles import Font
# Modules to Sort Files in Order
from natsort import natsorted

class excelFormat:     
            
    def xls2xlsx(self, excelFile, outputFolder):
        """
        Converts .xls Files to .xlsx Files That OpenPyxl Can Read
        If the File is Already a .xlsx Files, Do Nothing
        If the File is Neither a .xls Nor .xlsx, it Exits the Program
        """
        # Check That the Current Extension is .xls or .xlsx
        _, extension = os.path.splitext(excelFile)
        # If the Extension is .xlsx, the File is Ready; Do Nothing
        if extension == '.xlsx':
            return excelFile
        # If the Extension is Not .xls/.xlsx, Then the Data is in the Wrong Format; Exit Program
        if extension not in ['.xls', '.xlsx']:
            print("Cannot Convert File to .xlsx")
            sys.exit()
        
        # Create Output File Directory to Save Data ONLY If None Exists
        os.makedirs(outputFolder, exist_ok = True)
        # Convert '.xls' to '.xlsx'
        filename = os.path.basename(excelFile)
        newExcelFile = outputFolder + filename + "x"
        pyexcel.save_as(file_name = excelFile, dest_file_name = newExcelFile, logfile=open(os.devnull, 'w'))
        
        # Return New Excel File Name
        return newExcelFile
    
    def txt2csv(self, txtFile, csvFile, csvDelimiter = ",", overwriteCSV = False):
        # Check to See if CSV Conversion Alreayd Occurred
        if not os.path.isfile(csvFile) or overwriteCSV:
            with open(txtFile, "r") as inputData:
                in_reader = csv.reader(inputData, delimiter = csvDelimiter)
                with open(csvFile, 'w', newline='') as out_csv:
                    out_writer = csv.writer(out_csv)
                    for row in in_reader:
                        out_writer.writerow(row)
    
    def convertToExcel(self, inputFile, excelFile, excelDelimiter = ",", overwriteXL = False, testSheetNum = 0):
        """
        inputFile: The Input TXT/CSV File to Convert XLSX
        excelFile: The Output Excel File Name (XLSX)
        """
        # If the File is Not Already Converted: Convert the CSV to XLSX
        if not os.path.isfile(excelFile) or overwriteXL:
            # Make Excel WorkBook
            xlWorkbook = xl.Workbook()
            xlWorksheet = xlWorkbook.active
            # Write the Data from the CSV File to the Excel WorkBook
            with open(inputFile, "r") as inputData:
                inReader = csv.reader(inputData, delimiter = excelDelimiter)
                with open(excelFile, 'w+', newline=''):
                    for row in inReader:
                        xlWorksheet.append(row)
            # Save as New Excel File
            xlWorkbook.save(excelFile)
        # Else Load the GSR Data from the Excel File
        else:
            # Load the GSR Data from the Excel File
            xlWorkbook = xl.load_workbook(excelFile, data_only=True, read_only=True)
            xlWorksheet = xlWorkbook.worksheets[testSheetNum]
        
        # Return Excel Sheet
        return xlWorkbook, xlWorksheet

    def addExcelAesthetics(self, WB_worksheet):
        # Center the Data in the Cells
        align = Alignment(horizontal='center',vertical='center',wrap_text=True)        
        for column_cells in WB_worksheet.columns:
            length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
            WB_worksheet.column_dimensions[xl.utils.get_column_letter(column_cells[0].column)].width = length
            
            for cell in column_cells:
                cell.alignment = align
        # Header Style
        for cell in WB_worksheet["1:1"]:
            cell.font = Font(color='00FF0000', italic=True, bold=True)
        
        return WB_worksheet
    

class processFiles(excelFormat):
    
    def getFiles(self, dataDirectory, fileDoesntContain, fileContains):
        # If Using All the CSV Files in the Folder
        analysisFile = []; filesAdded = set();
        for fileName in os.listdir(dataDirectory):
            fileBase = os.path.splitext(fileName)[0]
            if fileName.endswith((".txt",'csv','xlsx')) and fileDoesntContain not in fileName and fileContains in fileName and fileBase not in filesAdded:
                analysisFile.append(fileName)
                filesAdded.add(fileBase)
        if len(analysisFile) == 0:
            print("No TXT/CSV/XLSX Files Found in the Data Folder:", dataDirectory)
            print("Found the Following Files:", os.listdir(dataDirectory))
            sys.exit()
        
        return natsorted(analysisFile)
    
    def getExcelFile(self, oldFile, outputFolder, testSheetNum = 0, excelDelimiter = ","):
        """
        --------------------------------------------------------------------------
        Input Variable Definitions:
            oldFile: The Path to the Excel File Containing the Data: txt, csv, xls, xlsx
            testSheetNum: An Integer Representing the Excel Worksheet (0-indexed) Order.
        --------------------------------------------------------------------------
        """
        # Check if File Exists
        if not os.path.exists(oldFile):
            sys.exit("The following Input File Does Not Exist:", oldFile)

        # Convert the TXT and CSV Files to XLSX
        if oldFile.endswith((".txt", ".csv")):
            # Extract Filename Information
            oldFileExtension = os.path.basename(oldFile)
            filename = os.path.splitext(oldFileExtension)[0]
            newFilePath = outputFolder + "Excel Files/"
            # Make Output Folder Directory if Not Already Created
            os.makedirs(newFilePath, exist_ok = True)

            # Convert CSV or TXT to XLSX
            excelFile = newFilePath + filename + ".xlsx"
            xlWorkbook, xlWorksheet = self.convertToExcel(oldFile, excelFile, excelDelimiter = excelDelimiter, overwriteXL = False, testSheetNum = testSheetNum)
        # If the File is Already an Excel File, Just Load the File
        elif oldFile.endswith(".xlsx"):
            excelFile = oldFile
            # Convert the XLS Files (Old Excel Format Files) to XLSX
            if excelFile.endswith(".xls"):
                excelFile = self.xls2xlsx(excelFile, outputFolder)
            # Load the GSR Data from the Excel File
            xlWorkbook = xl.load_workbook(excelFile, data_only=True, read_only=True)
            xlWorksheet = xlWorkbook.worksheets[testSheetNum]
        else:
            sys.exit("The Following File is Neither CSV, TXT, Nor XLSX:", excelFile)
        
        # Return the Final Worksheet
        print("Processing Data:", excelFile.split("/")[-1])
        return xlWorksheet, xlWorkbook
    
    
class saveData(excelFormat):
    def __init__(self):
        super().__init__()
        
        self.emptySheetName = "Empty Sheet"
    
    def getExcelDocument(self, excelFile, overwriteSave = False):
        # If the excel file you are saving already exists.
        if os.path.isfile(excelFile):
            # If You Want to Overwrite the Excel.
            if overwriteSave:
                print("\t\tDeleting Old Excel Workbook")
                os.remove(excelFile) 
            else:
                print("\t\tNot overwriting the file ... but your file already exists??")
            
        # If the File is Not Present: Create The Excel File
        if not os.path.isfile(excelFile):
            print("\t\tCreating New Excel Workbook")
            # Make Excel WorkBook
            WB = xl.Workbook()
            worksheet = WB.active 
            worksheet.title = self.emptySheetName
        else:
            print("\t\tExcel File Already Exists. Adding New Sheet to File")
            WB = xl.load_workbook(excelFile, read_only=False)
            worksheet = WB.create_sheet(self.emptySheetName)
        return WB, worksheet
    
    def saveDataCV(self, peakInfoHolder, saveDataFolder, saveExcelName, sheetName = "CV Analysis"):
        print("Saving the Data")
        # Create Output File Directory to Save Data: If Not Already Created
        os.makedirs(saveDataFolder, exist_ok=True)
        
        # Create Path to Save the Excel File
        excelFile = saveDataFolder + saveExcelName
        
        # Get the excel document.
        WB, worksheet = self.getExcelDocument(excelFile, overwriteSave = True)
            
        headers = ["Cycle Number"]
        # Add a header label for each peak
        peakTypes = ["Oxidation", "Reduction"]
        for reductiveScan in range(len(peakInfoHolder)):
            peakType = peakTypes[reductiveScan]
            for peakNum in range(len(peakInfoHolder[reductiveScan])):
                peakInfoString = peakType + " Peak " + str(peakNum)
                headers.extend([peakInfoString + " Potential (V)", peakInfoString + " Current (uAmps)", ""])
        worksheet.append(headers)
        
        # Organize and save the data
        for frameNum in range(len(peakInfoHolder[0][0])):
            frameData = [frameNum+1]
            for reductiveScan in range(len(peakInfoHolder)):
                for peakNum in range(len(peakInfoHolder[reductiveScan])):
                    
                    Ip = peakInfoHolder[reductiveScan][peakNum][frameNum][1]
                    Ep = peakInfoHolder[reductiveScan][peakNum][frameNum][0]
                    frameData.extend([Ep, Ip, ""])

            # Write the Data to Excel
            worksheet.append(frameData)
        
        # Add Excel Aesthetics
        worksheet = self.addExcelAesthetics(worksheet)    
            
        # Save as New Excel File
        WB.save(excelFile)
        WB.close()
        
        
        
        
        
        
        
        
        
        
        
        
    
    
    